"""
ingest_portos.py — Catálogo de portos públicos → OpenSearch (mr_portos_v001)
=============================================================================

1. Lê a curadoria ``backend/data/portos_brasil.csv`` (coordenadas de portão,
   aliases, cargas, AP).
2. Baixa os CSVs poligonais mais recentes do CKAN (dataset MTransp
   ``85954902-5ec2-4432-a6f9-99f3ce77f8d1`` — Poligonais dos Portos Públicos).
3. Monta ``geo_shape`` (Polygon / MultiPolygon) + centróide + área aproximada.
4. Faz bulk index com ``_id = codigo`` (PSV, PNG, …).

Pré-requisitos:
  - Índice criado: ``python -m scripts.setup_indices --index mr_portos_v001``
  - Variáveis OpenSearch no .env (mesmo padrão do MCP).

Uso (a partir da pasta ``backend/``)::

    PYTHONPATH=. python -m scripts.ingest_portos
    PYTHONPATH=. python -m scripts.ingest_portos --dry-run
    PYTHONPATH=. python -m scripts.ingest_portos --only-codigo PSV --verbose
    PYTHONPATH=. python -m scripts.ingest_portos --skip-ckan   # só CSV (sem polígono)
    PYTHONPATH=. python -m scripts.ingest_portos --embed       # embedding_nome (Azure)
    PYTHONPATH=. python -m scripts.ingest_portos --config-yaml data/portos_ckan_overrides.yaml

Overrides CKAN: ficheiro YAML (opcional) ``data/portos_ckan_overrides.yaml`` — ver comentários no ficheiro.

Requer ``shapely``; para ``--embed`` requer Azure OpenAI no .env (mesmo padrão do MCP).
"""
from __future__ import annotations

import csv
import io
import json
import logging
import math
import re
import sys
import unicodedata
from collections import defaultdict
from dataclasses import dataclass
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

import click
import httpx
import yaml
from opensearchpy import OpenSearch, helpers
from openai import AzureOpenAI
from openpyxl import load_workbook
from shapely.geometry import MultiPolygon, Polygon, mapping as shapely_mapping
from shapely.validation import explain_validity

# mesmo bootstrap que setup_indices
sys.path.insert(0, str(Path(__file__).parent.parent))

from mcp_servers.common.config import mcp_settings
from scripts.setup_indices import ALL_INDICES, create_index

log = logging.getLogger("ingest_portos")

INDEX = "mr_portos_v001"
CKAN_PACKAGE_ID = "85954902-5ec2-4432-a6f9-99f3ce77f8d1"
CKAN_API = "https://dados.transportes.gov.br/api/3/action/package_show"
USER_AGENT = "MineralRadar-ingest-portos/1.0 (+https://github.com/mineralradar)"

# Simplificação ~100 m em graus (mesma ordem de grandeza do bot_municipios).
DEFAULT_SIMPLIFY = 0.001

# Fallback se o YAML não existir ou não listar o código.
BUILTIN_CKAN_MATCH: dict[str, tuple[str, str]] = {
    "ITQ": ("Itaqui", "MA"),
    "SUA": ("Suape", "PE"),
    "ATU": ("Aratu", "BA"),
    "VDC": ("Vila do Conde", "PA"),
    "BRR": ("Barra do Riacho", "ES"),
    "FRN": ("Forno", "RJ"),
}


@dataclass(frozen=True)
class PortosYamlConfig:
    """Resultado da leitura do YAML de overrides (e defaults mesclados)."""

    ckan_match: dict[str, tuple[str, str]]
    ckan_package_id: str | None


def load_portos_yaml_config(
    path: Path,
    *,
    must_exist: bool,
) -> PortosYamlConfig:
    """
    Mescla ``BUILTIN_CKAN_MATCH`` com ``ckan_match`` do YAML (chaves do YAML ganham).
    ``ckan_package_id`` no YAML é opcional (None = usar constante CKAN_PACKAGE_ID).
    """
    merged = dict(BUILTIN_CKAN_MATCH)
    pkg: str | None = None
    if not path.is_file():
        if must_exist:
            log.error("Ficheiro de config não encontrado: %s", path)
            sys.exit(1)
        return PortosYamlConfig(ckan_match=merged, ckan_package_id=None)

    raw = yaml.safe_load(path.read_text(encoding="utf-8")) or {}
    if not isinstance(raw, dict):
        log.error("YAML inválido (esperado mapping na raiz): %s", path)
        sys.exit(1)

    pkg_raw = raw.get("ckan_package_id")
    if isinstance(pkg_raw, str) and pkg_raw.strip():
        pkg = pkg_raw.strip()

    cm = raw.get("ckan_match") or {}
    if cm and not isinstance(cm, dict):
        log.error("ckan_match deve ser um mapping em %s", path)
        sys.exit(1)
    for code, spec in cm.items():
        if spec is None:
            continue
        if isinstance(spec, dict):
            cidade = str(spec.get("cidade", "")).strip()
            uf = str(spec.get("uf", "")).strip().upper()
            if cidade and uf and len(uf) == 2:
                merged[str(code).strip().upper()] = (cidade, uf)
            else:
                log.warning("Ignorando ckan_match[%s]: cidade/uf inválidos", code)
        else:
            log.warning("Ignorando ckan_match[%s]: esperado objeto com cidade e uf", code)

    log.info("YAML %s: %d entradas ckan_match (após merge com builtins)", path.name, len(merged))
    return PortosYamlConfig(ckan_match=merged, ckan_package_id=pkg)


def _embedding_text_for_row(doc: dict[str, Any]) -> str:
    parts = [
        str(doc.get("nome", "")).strip(),
        str(doc.get("uf", "")).strip(),
        ", ".join(doc.get("cargas_principais") or []),
        str(doc.get("aliases", "")).strip(),
    ]
    s = " | ".join(p for p in parts if p)
    return s.strip()[:8000]


def sync_embed_texts(texts: list[str]) -> list[list[float] | None]:
    """Azure OpenAI embeddings (síncrono, sem Redis). Ordem preservada."""
    n = len(texts)
    out: list[list[float] | None] = [None] * n
    if not texts:
        return out
    if not mcp_settings.azure_openai_endpoint or not mcp_settings.azure_openai_api_key:
        log.warning("--embed: Azure OpenAI não configurado (endpoint/api_key) — a saltar embeddings")
        return out

    dim = mcp_settings.embedding_dimensions
    client = AzureOpenAI(
        api_version=mcp_settings.azure_openai_embedding_api_version,
        azure_endpoint=mcp_settings.azure_openai_endpoint,
        api_key=mcp_settings.azure_openai_api_key,
    )
    deployment = mcp_settings.embedding_deployment
    batch = 16
    for i in range(0, n, batch):
        chunk = texts[i : i + batch]
        try:
            resp = client.embeddings.create(input=chunk, model=deployment)
            for j, item in enumerate(resp.data):
                vec = item.embedding
                if len(vec) != dim:
                    log.warning(
                        "embedding dim=%s diferente de embedding_dimensions=%s — doc ignorado",
                        len(vec),
                        dim,
                    )
                    continue
                out[i + j] = vec
        except Exception as exc:  # noqa: BLE001
            log.error("Falha ao gerar embeddings no lote offset=%s: %s", i, exc)
    return out


def _strip_accents(s: str) -> str:
    nfkd = unicodedata.normalize("NFD", s)
    return "".join(c for c in nfkd if not unicodedata.combining(c))


def _ckan_group_key(city: str, uf: str) -> str:
    """Agrupa recursos CKAN (ex.: Itaguai / Itaguaí)."""
    c = _strip_accents(city).casefold().strip()
    return f"{c}-{uf.upper()}"


def _nome_normalizado(nome: str) -> str:
    return _strip_accents(nome).lower()


def _parse_br_float(raw: str) -> float:
    s = (raw or "").strip().replace(",", ".")
    return float(s)


def get_os_client(timeout: int = 120) -> OpenSearch:
    endpoint = mcp_settings.opensearch_endpoint or "http://localhost:9200"
    kwargs: dict[str, Any] = {
        "hosts": [endpoint],
        "use_ssl": mcp_settings.opensearch_use_ssl,
        "verify_certs": mcp_settings.opensearch_verify_certs,
        "timeout": timeout,
    }
    if mcp_settings.opensearch_user and mcp_settings.opensearch_password:
        kwargs["http_auth"] = (mcp_settings.opensearch_user, mcp_settings.opensearch_password)
    return OpenSearch(**kwargs)


def load_curated_csv(path: Path) -> list[dict[str, Any]]:
    """Lê portos_brasil.csv ignorando linhas de comentário (#)."""
    rows: list[dict[str, Any]] = []
    with path.open(encoding="utf-8") as f:
        for line in f:
            if line.startswith("#") or not line.strip():
                continue
            # primeira linha não comentário = header
            header_line = line
            break
        else:
            return rows
        rdr = csv.DictReader(io.StringIO(header_line + f.read()), delimiter=";")
        for row in rdr:
            if not row.get("codigo"):
                continue
            rows.append(row)
    return rows


def _resource_format_rank(url: str, fmt: str) -> int:
    """Maior = preferido. CSV vence XLSX no mesmo trimestre."""
    u = (url or "").lower()
    f = (fmt or "").upper()
    if u.endswith(".csv") or f == "CSV":
        return 2
    if u.endswith(".xlsx") or u.endswith(".xls") or f in ("XLSX", "XLS", "Microsoft Excel"):
        return 1
    return 0


def _pick_best_ckan_resource(candidates: list[tuple[int, str, str]]) -> tuple[int, str]:
    """Maior score; em empate, preferir CSV a XLSX."""
    score, url, _ = max(candidates, key=lambda c: (c[0], _resource_format_rank(c[1], c[2])))
    return score, url


def fetch_ckan_latest_urls(client: httpx.Client, package_id: str) -> dict[str, tuple[int, str]]:
    """
    Por grupo estável (cidade normalizada + UF), devolve (score_yyyymm, url)
    do recurso mais recente (CSV preferido ao XLSX no mesmo trimestre).
    """
    resp = client.get(CKAN_API, params={"id": package_id}, timeout=60.0)
    resp.raise_for_status()
    data = resp.json()
    if not data.get("success"):
        raise RuntimeError(f"CKAN error: {data}")
    resources = data["result"]["resources"]
    pat_mm = re.compile(
        r"^(.+)-([A-Z]{2})\s*-\s*Trimestral\s*-\s*(\d{2})-(\d{4})\s*$",
        re.UNICODE,
    )
    pat_slash = re.compile(
        r"^(.+)-([A-Z]{2})\s*-\s*Trimestral\s*-\s*(\d{2})/(\d{4})\s*$",
        re.UNICODE,
    )
    grouped: dict[str, list[tuple[int, str, str]]] = defaultdict(list)
    for r in resources:
        name = (r.get("name") or "").strip()
        m = pat_mm.match(name) or pat_slash.match(name)
        if not m:
            log.debug("ignorando recurso CKAN sem padrão trimestral: %s", name[:80])
            continue
        city, uf, mm, yy = m.group(1), m.group(2), int(m.group(3)), int(m.group(4))
        gkey = _ckan_group_key(city, uf)
        score = yy * 100 + mm
        url = (r.get("url") or "").strip()
        if not url:
            continue
        fmt = (r.get("format") or "").strip()
        grouped[gkey].append((score, url, fmt))
    best: dict[str, tuple[int, str]] = {}
    for gkey, lst in grouped.items():
        best[gkey] = _pick_best_ckan_resource(lst)
    return best


def _lookup_ckan_url(
    codigo: str,
    municipio: str,
    uf: str,
    ckan_best: dict[str, tuple[int, str]],
    ckan_match: dict[str, tuple[str, str]],
) -> tuple[int | None, str | None]:
    code_u = codigo.strip().upper()
    if code_u in ckan_match:
        city, uf_o = ckan_match[code_u]
        uf = uf_o
    else:
        city = municipio
    gkey = _ckan_group_key(city, uf)
    hit = ckan_best.get(gkey)
    if hit:
        return hit[0], hit[1]
    return None, None


def _approx_area_km2(poly: Polygon | MultiPolygon) -> float:
    """Área aproximada (km²) usando escala local no centróide — metadado, não legal."""
    c = poly.centroid
    lat = math.radians(c.y)
    mx = 111_320.0 * max(math.cos(lat), 0.01)
    my = 110_574.0
    return float(poly.area) * mx * my / 1e6


def _csv_text_to_rows(text: str) -> list[dict[str, Any]]:
    return list(csv.DictReader(io.StringIO(text), delimiter=";"))


def _read_poligonal_xlsx_rows(content: bytes) -> list[dict[str, str]]:
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        ws = wb.active
        it = iter(ws.iter_rows(values_only=True))
        header = next(it, None)
        if not header:
            return []
        keys: list[str] = []
        for i, h in enumerate(header):
            hs = "" if h is None else str(h).strip()
            keys.append(hs if hs else f"_col{i}")
        out: list[dict[str, str]] = []
        for row in it:
            if not row:
                continue
            if all(c is None or (isinstance(c, str) and not c.strip()) for c in row):
                continue
            d: dict[str, str] = {}
            for k, cell in zip(keys, row):
                if cell is None:
                    d[k] = ""
                elif isinstance(cell, (int, float)):
                    d[k] = str(cell)
                else:
                    d[k] = str(cell).strip()
            out.append(d)
        return out
    finally:
        wb.close()


def _row_anexo_lon_lat(row: dict[str, Any]) -> tuple[str, str, str] | None:
    """Extrai anexo, lon_raw, lat_raw (colunas tipo CSV MTransp / Excel)."""
    anexo = "1"
    lon_raw: str | None = None
    lat_raw: str | None = None
    for k, v in row.items():
        if v is None:
            continue
        vs = str(v).strip()
        if not vs:
            continue
        ks = str(k).strip()
        kl = ks.lower().replace(" ", "")
        if "utm" in kl:
            continue
        if re.match(r"^Anexos?$", ks, re.I) or (kl.startswith("anexo") and "vértice" not in kl):
            anexo = vs or "1"
        elif re.match(r"^Long\s*\(", ks, re.I):
            lon_raw = vs
        elif re.match(r"^Lat\s*\(", ks, re.I):
            lat_raw = vs
        elif kl == "longitude":
            lon_raw = lon_raw or vs
        elif kl == "latitude":
            lat_raw = lat_raw or vs
    if lon_raw is None or lat_raw is None:
        return None
    return anexo, lon_raw, lat_raw


def rows_vertex_to_geometry(
    rows: list[dict[str, Any]],
    simplify: float,
) -> tuple[dict[str, Any] | None, tuple[float, float] | None, float | None, str | None]:
    """
    Converte linhas tabulares (CSV ou Excel) em GeoJSON ``geo_shape``.
    """
    warnings: list[str] = []
    groups: dict[str, list[tuple[float, float]]] = defaultdict(list)
    for row in rows:
        parts = _row_anexo_lon_lat(row)
        if not parts:
            continue
        anexo, lon_raw, lat_raw = parts
        try:
            lon = _parse_br_float(lon_raw)
            lat = _parse_br_float(lat_raw)
        except ValueError:
            continue
        groups[anexo].append((lon, lat))

    def _anexo_sort_key(k: str) -> tuple[int, str]:
        return (int(k) if k.isdigit() else 9999, k)

    polys: list[Polygon] = []
    for anexo in sorted(groups.keys(), key=_anexo_sort_key):
        ring = groups[anexo]
        if len(ring) < 3:
            warnings.append(f"anexo {anexo}: <3 vértices")
            continue
        if ring[0] != ring[-1]:
            ring = [*ring, ring[0]]
        try:
            p = Polygon(ring)
            if not p.is_valid:
                p = p.buffer(0)
                if not p.is_valid:
                    warnings.append(f"anexo {anexo}: {explain_validity(p)[:120]}")
                    continue
            polys.append(p)
        except Exception as exc:  # noqa: BLE001
            warnings.append(f"anexo {anexo}: {exc}")

    if not polys:
        return None, None, None, "; ".join(warnings) if warnings else "sem polígonos válidos"

    geom: Polygon | MultiPolygon = polys[0] if len(polys) == 1 else MultiPolygon(polys)
    geom = geom.simplify(simplify, preserve_topology=True)
    if geom.is_empty:
        return None, None, None, "geometria vazia após simplify"

    c0 = geom.centroid
    if not geom.contains(c0):
        c0 = geom.representative_point()
    centroid = (c0.x, c0.y)

    gj = shapely_mapping(geom)
    area = _approx_area_km2(geom)
    msg = "; ".join(warnings) if warnings else None
    return gj, centroid, area, msg


def poligonal_bytes_to_geometry(
    content: bytes,
    url: str,
    simplify: float,
) -> tuple[dict[str, Any] | None, tuple[float, float] | None, float | None, str | None]:
    """Descodifica resposta CKAN (CSV ou XLSX) e monta geometria."""
    ul = (url or "").lower()
    try:
        if ul.endswith(".xlsx") or ul.endswith(".xls"):
            rows = _read_poligonal_xlsx_rows(content)
            if not rows:
                return None, None, None, "xlsx sem linhas legíveis"
        else:
            text = content.decode("utf-8-sig", errors="replace")
            rows = _csv_text_to_rows(text)
        return rows_vertex_to_geometry(rows, simplify)
    except Exception as exc:  # noqa: BLE001
        return None, None, None, f"erro ao ler poligonal: {exc}"


def row_to_os_doc(
    row: dict[str, str],
    poligono: dict[str, Any] | None,
    centroide_poly: tuple[float, float] | None,
    area_km2: float | None,
    ckan_score: int | None,
) -> dict[str, Any]:
    cargas = [x.strip() for x in (row.get("cargas_principais") or "").split(",") if x.strip()]
    aliases = (row.get("aliases") or "").strip()
    lat = float(row["latitude"])
    lon = float(row["longitude"])
    vp = (row.get("validacao_pendente") or "false").strip().lower() in ("1", "true", "yes", "sim")

    doc: dict[str, Any] = {
        "codigo": row["codigo"].strip(),
        "nome": row["nome"].strip(),
        "nome_normalizado": _nome_normalizado(row["nome"]),
        "tipo": "PORTO_ORGANIZADO",
        "esfera": "FEDERAL",
        "uf": row["uf"].strip().upper(),
        "municipio": row["municipio"].strip(),
        "autoridade_portuaria": (row.get("autoridade_portuaria") or "").strip(),
        "endereco": (row.get("endereco") or "").strip(),
        "cargas_principais": cargas,
        "centroide": (
            {"lat": round(centroide_poly[1], 6), "lon": round(centroide_poly[0], 6)}
            if centroide_poly
            else {"lat": lat, "lon": lon}
        ),
        "acesso_rodoviario": {"lat": lat, "lon": lon},
        "aliases": aliases,
        "fonte": "mtransp_ckan_poligonais+portos_brasil_csv",
        "validacao_pendente": vp,
        "ativo": True,
        "indexed_at": datetime.now(timezone.utc).isoformat(),
    }
    if poligono is not None:
        doc["poligono"] = poligono
    if area_km2 is not None:
        doc["area_km2"] = round(area_km2, 3)
    if ckan_score is not None:
        yy, mm = divmod(ckan_score, 100)
        doc["data_referencia"] = f"{yy:04d}-{mm:02d}-01"
    return doc


@click.command()
@click.option("--dry-run", is_flag=True, help="Não grava no OpenSearch.")
@click.option("--verbose", "-v", is_flag=True, help="Log DEBUG.")
@click.option("--skip-ckan", is_flag=True, help="Só CSV curado (sem baixar poligonais).")
@click.option("--recreate-index", is_flag=True, help="Apaga e recria mr_portos_v001 (DESTRUTIVO).")
@click.option("--only-codigo", default=None, help="Indexa apenas um código (ex.: PSV).")
@click.option("--simplify", default=DEFAULT_SIMPLIFY, type=float, show_default=True, help="Tolerância shapely.simplify (graus).")
@click.option(
    "--csv-path",
    type=click.Path(path_type=Path),
    default=None,
    help="Override do caminho para portos_brasil.csv",
)
@click.option(
    "--config-yaml",
    "config_yaml",
    type=click.Path(path_type=Path),
    default=None,
    help="YAML (ckan_match, ckan_package_id). Omisso: data/portos_ckan_overrides.yaml se existir.",
)
@click.option(
    "--package-id",
    default=None,
    help="UUID do pacote CKAN poligonal (prevalece sobre YAML e default).",
)
@click.option(
    "--embed",
    is_flag=True,
    default=False,
    help="Gera embedding_nome via Azure OpenAI (síncrono; sem Redis).",
)
def main(
    dry_run: bool,
    verbose: bool,
    skip_ckan: bool,
    recreate_index: bool,
    only_codigo: str | None,
    simplify: float,
    csv_path: Path | None,
    config_yaml: Path | None,
    package_id: str | None,
    embed: bool,
) -> None:
    logging.basicConfig(
        level=logging.DEBUG if verbose else logging.INFO,
        format="%(levelname)s  %(message)s",
    )
    backend_root = Path(__file__).resolve().parents[1]
    csv_file = csv_path or (backend_root / "data" / "portos_brasil.csv")
    if not csv_file.is_file():
        log.error("CSV não encontrado: %s", csv_file)
        sys.exit(1)

    default_yaml = backend_root / "data" / "portos_ckan_overrides.yaml"
    if config_yaml is None:
        yaml_path = default_yaml
        yaml_must_exist = False
    else:
        yaml_path = Path(config_yaml)
        yaml_must_exist = True

    yaml_cfg = load_portos_yaml_config(yaml_path, must_exist=yaml_must_exist)
    if config_yaml is None and not yaml_path.is_file():
        log.info("Sem %s — a usar só overrides embutidos no script", yaml_path.name)

    ckan_match = yaml_cfg.ckan_match
    package_effective = (package_id or yaml_cfg.ckan_package_id or CKAN_PACKAGE_ID).strip()
    if package_effective != CKAN_PACKAGE_ID:
        log.info("CKAN package_id=%s", package_effective)

    curated = load_curated_csv(csv_file)
    if only_codigo:
        curated = [r for r in curated if r["codigo"].strip().upper() == only_codigo.strip().upper()]
        if not curated:
            log.error("Nenhuma linha para codigo=%s", only_codigo)
            sys.exit(1)

    ckan_best: dict[str, tuple[int, str]] = {}
    if not skip_ckan:
        with httpx.Client(headers={"User-Agent": USER_AGENT}, follow_redirects=True) as h:
            ckan_best = fetch_ckan_latest_urls(h, package_effective)
        log.info("CKAN: %d grupos de porto com URL mais recente", len(ckan_best))

    os_client: OpenSearch | None = None
    if not dry_run:
        os_client = get_os_client()
        body = ALL_INDICES[INDEX]["body"]
        create_index(os_client, INDEX, body, recreate=recreate_index)

    actions: list[dict[str, Any]] = []
    with httpx.Client(headers={"User-Agent": USER_AGENT}, follow_redirects=True, timeout=120.0) as http:
        for row in curated:
            codigo = row["codigo"].strip()
            poligono = None
            centroide_poly = None
            area_km2 = None
            ckan_score: int | None = None
            warn: str | None = None

            if not skip_ckan:
                score, url = _lookup_ckan_url(
                    codigo, row["municipio"], row["uf"], ckan_best, ckan_match
                )
                if url:
                    ckan_score = score
                    try:
                        r = http.get(url)
                        r.raise_for_status()
                        poligono, centroide_poly, area_km2, warn = poligonal_bytes_to_geometry(
                            r.content, url, simplify=simplify
                        )
                        if warn:
                            log.warning("%s: %s", codigo, warn)
                    except httpx.HTTPError as exc:
                        log.warning("%s: falha HTTP ao baixar poligonal: %s", codigo, exc)
                else:
                    log.warning(
                        "%s (%s/%s): sem URL CKAN — documento só com ponto curado",
                        codigo,
                        row.get("municipio"),
                        row.get("uf"),
                    )

            doc = row_to_os_doc(row, poligono, centroide_poly, area_km2, ckan_score)
            actions.append(
                {
                    "_op_type": "index",
                    "_index": INDEX,
                    "_id": codigo,
                    "_source": doc,
                }
            )
            log.info("preparado %s poligono=%s", codigo, poligono is not None)

    if embed:
        texts = [_embedding_text_for_row(a["_source"]) for a in actions]
        vecs = sync_embed_texts(texts)
        n_emb = 0
        for action, vec in zip(actions, vecs):
            if vec is not None:
                action["_source"]["embedding_nome"] = vec
                n_emb += 1
        log.info("embeddings: %d/%d documentos com vetor", n_emb, len(actions))

    if dry_run:
        log.info("dry-run: %d documentos (não enviados)", len(actions))
        if verbose and actions:
            log.debug(json.dumps(actions[0]["_source"], ensure_ascii=False, indent=2)[:2000])
        return

    assert os_client is not None
    success, failed = helpers.bulk(
        os_client,
        actions,
        refresh="wait_for",
        raise_on_error=False,
    )
    if failed:
        log.error("bulk com falhas: ok=%s primeiros_erros=%s", success, failed[:2])
        sys.exit(1)
    log.info("bulk concluído: %d documentos em %s", success, INDEX)


if __name__ == "__main__":
    main()
